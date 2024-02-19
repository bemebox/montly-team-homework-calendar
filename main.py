import calendar
from collections import namedtuple
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from web_client import get_request, ApiError

# create constant values
Constants = namedtuple("Constants", ["public_holidays_base_url", "country_code"])
constants = Constants("https://date.nager.at/api/v3/PublicHolidays", "PT")


def main():
    # get current date
    current_date = get_current_date()

    # get current month country public holidays
    month_public_holidays_by_month = get_country_public_holidays_by_month(
        constants.country_code, current_date.year, current_date.month
    )

    # create the Excel file
    excel_file = create_calendar_excel(
        current_date.year, current_date.month, month_public_holidays_by_month
    )

    # save the Excel file
    excel_file.save(f"calendar_{current_date.year}_{current_date.month}.xlsx")


def get_current_date():
    """
    function that get the current date.

    Returns:
        datetime.date: the current date.
    """
    # get the current date and time
    current_date_time = datetime.now()

    # return the current date
    return current_date_time.date()


def get_country_public_holidays_by_month(country_code, year, month):
    """
    get public holidays for a specific country, year, and month.

    Args:
        country_code (str): the ISO country code (e.g., 'US', 'GB').
        year (int): the year for which to retrieve public holidays.
        month (int): the month for which to retrieve public holidays (1 to 12).

    Returns:
        list: a list of public holidays for the specified country, year, and month.
              each holiday is represented as a dictionary with date, year, month, day, and description.

    Example:
        get_country_public_holidays_by_month('US', 2024, 7)
    """
    month_holidays = []

    year_country_public_holidays = get_country_public_holidays_by_year(
        country_code, year
    )
    if year_country_public_holidays:
        for holiday in year_country_public_holidays:
            public_holiday_date = datetime.strptime(holiday["date"], "%Y-%m-%d")
            if public_holiday_date.month == month:
                month_holidays.append(
                    {
                        "date": holiday["date"],
                        "year": public_holiday_date.year,
                        "month": public_holiday_date.month,
                        "day": public_holiday_date.day,
                        "description": holiday["localName"],
                    }
                )

    return month_holidays


def get_country_public_holidays_by_year(country_code, year):
    """
    get public holidays for a specific country and year, from an external API.

    Args:
        country_code (str): the ISO country code (e.g., 'US', 'GB').
        year (int): the year for which to retrieve public holidays.

    Returns:
        list or None: a list of public holidays for the specified country and year.
                      returns None if the request fails.

    Example:
        get_country_public_holidays_by_year('US', 2024)
    """
    endpoint = f"{constants.public_holidays_base_url}/{year}/{country_code}"

    try:
        return get_request(endpoint)
    except ApiError as ex:
        print(f"API Error: {ex}")
        return None


def create_calendar_excel(year, month, holidays=[]):
    """
    create an Excel workbook with a calendar for a specific year and month.

    Args:
        year (int): the year for which to create the calendar.
        month (int): the month for which to create the calendar (1 to 12).
        holidays (list, optional): list of holiday dates (format: 'YYYY-MM-DD'). Defaults to an empty list.

    Returns:
        openpyxl.workbook.Workbook: the Excel workbook containing the calendar.

    Example:
        create_calendar_excel(2024, 2, ["2024-02-14", "2024-02-21"])
    """
    # create a new workbook
    workbook = Workbook()
    sheet = workbook.active

    # initialize the workbook sheet rows
    sheet.cell(row=1, column=1, value="")
    sheet.cell(row=2, column=1, value="")

    # get the days of the month's calendar
    month_calendar = calendar.monthcalendar(year, month)

    # populate the Excel sheet with days and mark holidays in gray
    for week in month_calendar:
        for day in week:
            if day != 0:

                # add the day of the week in the first row
                weekday = calendar.day_abbr[calendar.weekday(year, month, day)]
                sheet.cell(row=1, column=sheet.max_column + 1, value=weekday)

                # add the day of the month in the second row
                weekday = calendar.day_abbr[calendar.weekday(year, month, day)]
                sheet.cell(row=2, column=sheet.max_column, value=day)

                date = f"{year}-{month:02d}-{day:02d}"
                if date in holidays:
                    # Fill the cell with gray color for holidays
                    sheet.cell(row=sheet.max_row, column=sheet.max_column).fill = (
                        PatternFill(
                            start_color="808080", end_color="808080", fill_type="solid"
                        )
                    )

    return workbook


if __name__ == "__main__":
    main()
